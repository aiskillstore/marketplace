#!/usr/bin/env python3
"""Validate OT workbook structure and compare calculated values."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from recompute_expected_values import InputError, calculate, load_payload


REQUIRED_SHEETS = [
    "OT Cost Summary",
    "Milestone Detail",
    "Scenario Analysis",
    "Labor Benchmarking",
    "Cost Share & Funding",
    "Methodology",
    "Raw Data",
]
CELL_REF = re.compile(r"^(?:'((?:[^']|'')+)'|([^!]+))!\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")


def normalize_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value)).upper()


def parse_cell_ref(reference: str) -> tuple[str, str]:
    match = CELL_REF.fullmatch(reference.strip())
    if not match:
        raise InputError(f"invalid workbook cell reference: {reference}")
    sheet = (match.group(1) or match.group(2)).replace("''", "'")
    return sheet, f"{match.group(3).upper()}{match.group(4)}"


def value_at(workbook: Any, reference: str) -> Any:
    sheet, coordinate = parse_cell_ref(reference)
    if sheet not in workbook.sheetnames:
        raise InputError(f"cell reference uses missing sheet: {reference}")
    return workbook[sheet][coordinate].value


def check_formula(
    failures: list[str],
    workbook: Any,
    reference: str,
    *,
    expected: str | None = None,
    contains: list[str] | None = None,
    not_contains: list[str] | None = None,
) -> None:
    try:
        value = value_at(workbook, reference)
    except InputError as exc:
        failures.append(str(exc))
        return
    if not isinstance(value, str) or not value.startswith("="):
        failures.append(f"{reference} is not a formula")
        return
    normalized = normalize_formula(value)
    if expected is not None and normalized != normalize_formula(expected):
        failures.append(f"{reference} formula does not match expected structure")
    for item in contains or []:
        if normalize_formula(item) not in normalized:
            failures.append(f"{reference} formula is missing {item}")
    for item in not_contains or []:
        if normalize_formula(item) in normalized:
            failures.append(f"{reference} formula contains forbidden text {item}")


def structural_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    for sheet_name in payload.get("required_sheets", REQUIRED_SHEETS):
        if sheet_name not in workbook.sheetnames:
            failures.append(f"missing required sheet: {sheet_name}")

    if "OT Cost Summary" in workbook.sheetnames:
        check_formula(
            failures,
            workbook,
            "'OT Cost Summary'!B8",
            contains=["VALUE(LEFT(B7,4))", "VALUE(LEFT(B6,4))", "VALUE(MID(B7,6,2))", "VALUE(MID(B6,6,2))"],
            not_contains=["DATEDIF", "YEAR("],
        )
        check_formula(
            failures,
            workbook,
            "'OT Cost Summary'!B9",
            contains=["B3", "B8", "^"],
        )
        summary = workbook["OT Cost Summary"]
        if summary["B9"].number_format != "0.0000":
            failures.append("OT Cost Summary!B9 must display the aging factor as 0.0000")

    formula_count = 0
    all_text: list[str] = []
    error_tokens = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    all_text.append(value)
                    if value.startswith("="):
                        formula_count += 1
                        if any(token in value.upper() for token in error_tokens):
                            failures.append(f"{sheet.title}!{cell.coordinate} contains a formula error token")
                    elif value[:1] in {"+", "-", "@"}:
                        failures.append(f"{sheet.title}!{cell.coordinate} starts with a formula-trigger character")
    if formula_count == 0:
        failures.append("workbook contains no formulas")

    joined = "\n".join(all_text)
    if re.search(r"4022\s*\(d\)\s*\(1\)\s*\(D\).{0,80}competition commitment", joined, re.I | re.S):
        failures.append("workbook misstates 4022(d)(1)(D) as competition commitment")
    if re.search(r"4021.{0,100}(?:100%|fully)\s+Government funded", joined, re.I | re.S):
        failures.append("workbook automatically states 4021 is fully Government funded")
    if re.search(r"4022\s*\(f\).{0,100}(?:100%|fully)\s+Government funded", joined, re.I | re.S):
        failures.append("workbook automatically states 4022(f) is fully Government funded")

    assertions = payload.get("formula_assertions", [])
    if not isinstance(assertions, list):
        raise InputError("formula_assertions must be an array")
    for index, assertion in enumerate(assertions):
        if not isinstance(assertion, dict) or not isinstance(assertion.get("cell"), str):
            raise InputError(f"formula_assertions[{index}] must contain a cell string")
        contains = assertion.get("contains", [])
        not_contains = assertion.get("not_contains", [])
        if not isinstance(contains, list) or not all(isinstance(item, str) for item in contains):
            raise InputError(f"formula_assertions[{index}].contains must be strings")
        if not isinstance(not_contains, list) or not all(isinstance(item, str) for item in not_contains):
            raise InputError(f"formula_assertions[{index}].not_contains must be strings")
        expected = assertion.get("equals")
        if expected is not None and not isinstance(expected, str):
            raise InputError(f"formula_assertions[{index}].equals must be a string")
        check_formula(
            failures,
            workbook,
            assertion["cell"],
            expected=expected,
            contains=contains,
            not_contains=not_contains,
        )
    return failures


def find_soffice() -> Path | None:
    command = shutil.which("soffice")
    if command:
        return Path(command)
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return mac if mac.is_file() else None


def recalculate(source: Path, executable: Path) -> tuple[tempfile.TemporaryDirectory[str], Path]:
    temporary = tempfile.TemporaryDirectory(prefix="ot-workbook-validation-")
    root = Path(temporary.name)
    input_dir = root / "input"
    output_dir = root / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    copied = input_dir / source.name
    shutil.copy2(source, copied)
    completed = subprocess.run(
        [str(executable), "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(copied)],
        capture_output=True,
        text=True,
        timeout=120,
        check=False,
    )
    result = output_dir / source.name
    if completed.returncode != 0 or not result.is_file():
        temporary.cleanup()
        detail = (completed.stderr or completed.stdout).strip()
        raise InputError(f"LibreOffice calculation failed: {detail or 'no output file'}")
    return temporary, result


def close_enough(actual: float, expected: float, tolerance: float) -> bool:
    return math.isclose(actual, expected, rel_tol=tolerance, abs_tol=max(0.01, tolerance))


def compare(workbook: Any, expected: dict[str, Any], tolerance: float) -> list[str]:
    failures: list[str] = []

    def one(reference: str, target: float, label: str) -> None:
        try:
            value = value_at(workbook, reference)
        except InputError as exc:
            failures.append(str(exc))
            return
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            failures.append(f"{label} at {reference} has no calculated numeric value")
        elif not close_enough(float(value), target, tolerance):
            failures.append(f"{label} at {reference} is {float(value):.6f}, expected {target:.6f}")

    for milestone in expected["milestones"]:
        mapping = {
            "workbook_project_cost_cell": ("project_cost", "project cost"),
            "workbook_government_funding_cell": ("government_funding", "Government funding"),
            "workbook_ceiling_cell": ("ceiling_basis", "ceiling basis"),
            "workbook_performer_share_cell": ("performer_project_share", "performer share"),
        }
        for reference_key, (value_key, label) in mapping.items():
            if reference_key in milestone:
                one(milestone[reference_key], milestone[value_key], f"{milestone['id']} {label}")
    if "workbook_total_project_cost_cell" in expected:
        one(expected["workbook_total_project_cost_cell"], expected["total_project_cost"], "total project cost")
    if "workbook_total_government_funding_cell" in expected:
        one(
            expected["workbook_total_government_funding_cell"],
            expected["total_government_funding"],
            "total Government funding",
        )
    return failures


def cached_error_audit(workbook: Any) -> list[str]:
    failures: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    failures.append(f"{sheet.title}!{cell.coordinate} has cached error {value}")
    return failures


def run(path: Path, expected_path: Path, engine: str, tolerance: float) -> dict[str, Any]:
    if not zipfile.is_zipfile(path):
        return {"status": "fail", "failures": ["file is not a valid XLSX ZIP"]}
    payload = load_payload(expected_path)
    expected = calculate(payload)
    workbook = load_workbook(path, data_only=False)
    failures = structural_audit(workbook, payload)
    engine_used = "none"
    temporary: tempfile.TemporaryDirectory[str] | None = None
    try:
        executable = find_soffice() if engine in {"auto", "libreoffice"} else None
        if engine == "libreoffice" and executable is None:
            failures.append("LibreOffice was required but no executable was found")
        elif executable is not None:
            try:
                temporary, calculated_path = recalculate(path, executable)
                calculated = load_workbook(calculated_path, data_only=True)
                failures.extend(cached_error_audit(calculated))
                failures.extend(compare(calculated, expected, tolerance))
                engine_used = "libreoffice"
            except InputError as exc:
                failures.append(str(exc))
    finally:
        if temporary is not None:
            temporary.cleanup()
    return {
        "status": "pass" if not failures else "fail",
        "engine": engine_used,
        "formula_count": sum(
            1
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ),
        "failures": failures,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate an OT Cost Analysis workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--expected", required=True, type=Path)
    parser.add_argument("--engine", choices=("none", "auto", "libreoffice"), default="auto")
    parser.add_argument("--tolerance", type=float, default=0.01)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if not args.workbook.is_file() or not args.expected.is_file():
        print("ERROR: workbook or expected-input file not found", file=sys.stderr)
        return 2
    try:
        result = run(args.workbook, args.expected, args.engine, args.tolerance)
    except (InputError, OSError, ValueError, zipfile.BadZipFile) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    if args.json:
        print(json.dumps(result, indent=2, sort_keys=True))
    elif result["status"] == "pass":
        print(f"OT workbook validation passed; formula engine: {result['engine']}.")
    else:
        print("VALIDATION FAILED")
        for failure in result["failures"]:
            print(f"- {failure}")
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
