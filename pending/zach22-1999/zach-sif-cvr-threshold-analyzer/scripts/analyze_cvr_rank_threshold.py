#!/usr/bin/env python3
"""Analyze CVR warning thresholds against SIF organic-rank movement.

Public edition: this script only reads local business reports and local SIF
raw JSON exported by the user's own MCP/tooling. It never calls a private SIF
endpoint directly.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import unescape
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - CSV inputs still work without openpyxl.
    load_workbook = None


DEFAULT_SITE = "US"
DEFAULT_ANALYZER_DIR = Path("outputs") / "zach-sif-cvr-threshold-analyzer"
P1_MAX_RANK = 48


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "day": ("时间", "日期", "date", "day", "report date", "report_date"),
    "asin": ("asin", "ASIN", "子ASIN", "商品ASIN"),
    "sessions_total": (
        "Sessions-Total",
        "Sessions Total",
        "sessions-total",
        "sessions_total",
        "sessions",
        "session",
        "会话数",
        "总会话数",
        "访客数",
        "父体访客数",
    ),
    "cvr": ("CVR", "cvr", "Conversion Rate", "conversion rate", "转化率", "整体CVR", "买家转化率"),
    "ad_cvr": ("广告CVR", "ad_cvr", "Ad CVR", "Advertising CVR", "PPC CVR", "广告转化率"),
    "organic_cvr": ("自然CVR", "organic_cvr", "Organic CVR", "Natural CVR", "自然转化率"),
    "clicks": ("点击", "clicks", "Clicks", "广告点击", "广告点击量"),
    "ad_orders": ("广告订单量", "ad_orders", "Ad Orders", "Advertising Orders", "广告订单"),
    "organic_clicks": ("自然点击量", "organic_clicks", "Organic Clicks", "自然点击"),
    "organic_orders": ("自然订单量", "organic_orders", "Organic Orders", "自然订单"),
    "units": ("销量", "units", "Units", "Units Ordered", "销售量", "商品销量"),
    "orders": ("订单量", "orders", "Orders", "Total Orders", "订单"),
}


@dataclass
class BusinessRow:
    day: date
    sessions_total: int | None = None
    cvr: float | None = None
    ad_cvr: float | None = None
    organic_cvr: float | None = None
    clicks: int | None = None
    ad_orders: int | None = None
    organic_clicks: int | None = None
    organic_orders: int | None = None
    units: int | None = None
    orders: int | None = None
    raw: dict[str, Any] | None = None


@dataclass
class AnalysisConfig:
    business_report: Path
    asin: str
    site: str
    brand: str
    launch_date: date | None
    core_keywords: list[str]
    sif_cache: Path | None
    output_dir: Path
    analysis_date: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--business-report", required=True, type=Path, help="Daily business report, .csv/.xlsx/.xlsm.")
    parser.add_argument("--asin", required=True, help="Target ASIN.")
    parser.add_argument("--site", default=DEFAULT_SITE, help="Amazon site code, default: US.")
    parser.add_argument("--brand", default="UnknownBrand", help="Brand name for output grouping.")
    parser.add_argument("--launch-date", help="Offsite scaling start date, YYYY-MM-DD.")
    parser.add_argument("--core-keywords", help="Comma-separated core keywords. If omitted, infer from SIF score fields.")
    parser.add_argument("--sif-cache", type=Path, help="Local SIF daily keyword raw JSON from the user's own MCP/tooling.")
    parser.add_argument("--output-dir", type=Path, help="Output directory. Defaults to outputs/zach-sif-cvr-threshold-analyzer/{brand}.")
    parser.add_argument("--analysis-date", help="Output filename date, default: today.")
    return parser.parse_args()


def build_config(args: argparse.Namespace) -> AnalysisConfig:
    brand = (args.brand or "UnknownBrand").strip()
    output_dir = args.output_dir or DEFAULT_ANALYZER_DIR / safe_path_part(brand)
    return AnalysisConfig(
        business_report=args.business_report,
        asin=args.asin.strip().upper(),
        site=(args.site or DEFAULT_SITE).strip().upper(),
        brand=brand,
        launch_date=parse_optional_day(args.launch_date),
        core_keywords=split_keywords(args.core_keywords),
        sif_cache=args.sif_cache,
        output_dir=output_dir,
        analysis_date=args.analysis_date or date.today().isoformat(),
    )


def split_keywords(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip().lower() for item in value.split(",") if item.strip()]


def safe_path_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("_") or "UnknownBrand"


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def build_header_index(headers: list[Any]) -> dict[str, int]:
    normalized = {normalize_header(header): idx for idx, header in enumerate(headers)}
    result: dict[str, int] = {}
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            idx = normalized.get(normalize_header(alias))
            if idx is not None:
                result[canonical] = idx
                break
    return result


def parse_optional_day(value: Any) -> date | None:
    if value in (None, ""):
        return None
    return parse_day(value)


def parse_day(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError as exc:
        raise ValueError(f"Unsupported date value: {value!r}") from exc


def parse_percent(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        number = float(value)
        return number / 100 if abs(number) > 1 else number
    text = str(value).strip()
    if not text or text in {"-", "--", "—"}:
        return None
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1]
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return None
    return number / 100 if is_percent or abs(number) > 1 else number


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--", "—"}:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def fmt_pct(value: float | None, digits: int = 2) -> str:
    if value is None:
        return ""
    return f"{value * 100:.{digits}f}%"


def fmt_num(value: float | int | None, digits: int = 2) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    return f"{value:.{digits}f}"


def load_business_rows(config: AnalysisConfig) -> list[BusinessRow]:
    suffix = config.business_report.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        raw_rows = read_xlsx(config.business_report)
    elif suffix == ".csv":
        raw_rows = read_csv(config.business_report)
    else:
        raise RuntimeError(f"Unsupported business report format: {config.business_report}")

    rows = parse_business_raw_rows(raw_rows, config.asin)
    rows.sort(key=lambda row: row.day)
    if not rows:
        raise RuntimeError("No business report rows found for the target ASIN.")
    validate_date_continuity(rows)
    return rows


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    if load_workbook is not None:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = list(rows[0])
            return [{str(headers[i] or ""): value for i, value in enumerate(values)} for values in rows[1:]]
        except Exception:
            return read_xlsx_minimal(path)
    return read_xlsx_minimal(path)


def read_xlsx_minimal(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in archive.namelist():
            sheet_candidates = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet"))
            if not sheet_candidates:
                return []
            sheet_name = sheet_candidates[0]
        sheet_xml = archive.read(sheet_name).decode("utf-8", errors="replace")

    table: list[list[Any]] = []
    for row_xml in re.findall(r"<(?:[A-Za-z0-9_]+:)?row\b[^>]*>(.*?)</(?:[A-Za-z0-9_]+:)?row>", sheet_xml, flags=re.S):
        row_values: list[Any] = []
        for attrs_text, cell_xml in re.findall(
            r"<(?:[A-Za-z0-9_]+:)?c\b([^>]*)>(.*?)</(?:[A-Za-z0-9_]+:)?c>",
            row_xml,
            flags=re.S,
        ):
            attrs = parse_xml_attrs(attrs_text)
            ref = attrs.get("r", "")
            col_index = column_index_from_cell_ref(ref)
            while len(row_values) <= col_index:
                row_values.append(None)
            row_values[col_index] = parse_xlsx_cell(attrs, cell_xml, shared_strings)
        table.append(row_values)
    if not table:
        return []
    headers = [str(value or "") for value in table[0]]
    return [{headers[i]: values[i] if i < len(values) else None for i in range(len(headers))} for values in table[1:]]


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    shared_xml = archive.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    strings: list[str] = []
    for item_xml in re.findall(r"<(?:[A-Za-z0-9_]+:)?si\b[^>]*>(.*?)</(?:[A-Za-z0-9_]+:)?si>", shared_xml, flags=re.S):
        strings.append(extract_text_nodes(item_xml))
    return strings


def parse_xml_attrs(attrs_text: str) -> dict[str, str]:
    return {name: unescape(value) for name, value in re.findall(r'([A-Za-z_:][A-Za-z0-9_:.-]*)="([^"]*)"', attrs_text)}


def column_index_from_cell_ref(ref: str) -> int:
    letters = "".join(char for char in ref if char.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)


def parse_xlsx_cell(attrs: dict[str, str], cell_xml: str, shared_strings: list[str]) -> Any:
    cell_type = attrs.get("t")
    if cell_type == "inlineStr":
        return extract_text_nodes(cell_xml)
    value_match = re.search(r"<(?:[A-Za-z0-9_]+:)?v\b[^>]*>(.*?)</(?:[A-Za-z0-9_]+:)?v>", cell_xml, flags=re.S)
    if not value_match:
        return None
    text = unescape(value_match.group(1).strip())
    if cell_type == "s":
        idx = as_int(text)
        return shared_strings[idx] if idx is not None and 0 <= idx < len(shared_strings) else text
    if cell_type in {"str", "b"}:
        return text
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def extract_text_nodes(xml_text: str) -> str:
    texts = re.findall(r"<(?:[A-Za-z0-9_]+:)?t\b[^>]*>(.*?)</(?:[A-Za-z0-9_]+:)?t>", xml_text, flags=re.S)
    return "".join(unescape(text) for text in texts)


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def parse_business_raw_rows(raw_rows: list[dict[str, Any]], asin: str) -> list[BusinessRow]:
    if not raw_rows:
        return []
    headers = list(raw_rows[0].keys())
    index = build_header_index(headers)
    required = ["day", "sessions_total", "cvr"]
    missing = [field for field in required if field not in index]
    if missing:
        raise RuntimeError(f"Missing required business report columns: {missing}")

    rows: list[BusinessRow] = []
    for raw in raw_rows:
        values = list(raw.values())
        if "asin" in index:
            row_asin = str(values[index["asin"]] or "").strip().upper()
            if row_asin and row_asin != asin:
                continue
        rows.append(
            BusinessRow(
                day=parse_day(values[index["day"]]),
                sessions_total=as_int(values[index["sessions_total"]]),
                cvr=parse_percent(values[index["cvr"]]),
                ad_cvr=parse_percent(values[index["ad_cvr"]]) if "ad_cvr" in index else None,
                organic_cvr=parse_percent(values[index["organic_cvr"]]) if "organic_cvr" in index else None,
                clicks=as_int(values[index["clicks"]]) if "clicks" in index else None,
                ad_orders=as_int(values[index["ad_orders"]]) if "ad_orders" in index else None,
                organic_clicks=as_int(values[index["organic_clicks"]]) if "organic_clicks" in index else None,
                organic_orders=as_int(values[index["organic_orders"]]) if "organic_orders" in index else None,
                units=as_int(values[index["units"]]) if "units" in index else None,
                orders=as_int(values[index["orders"]]) if "orders" in index else None,
                raw=raw,
            )
        )
    return rows


def validate_date_continuity(rows: list[BusinessRow]) -> None:
    expected = rows[0].day
    gaps: list[str] = []
    for row in rows:
        if row.day != expected:
            gaps.append(f"{expected.isoformat()}..{row.day.isoformat()}")
            expected = row.day
        expected += timedelta(days=1)
    if gaps:
        raise RuntimeError(f"Business report date range is not continuous: {', '.join(gaps[:5])}")


def default_sif_cache_path(config: AnalysisConfig) -> Path:
    return config.output_dir / f"{config.analysis_date}_{config.site}_{config.asin}_sif-daily-keyword-raw.json"


def load_sif_cache(config: AnalysisConfig) -> dict[str, Any]:
    raw_path = config.sif_cache or default_sif_cache_path(config)
    if not raw_path.exists():
        raise RuntimeError(
            "Missing SIF cache JSON. Please use the user's own SIF MCP/tooling to export daily keyword rank data, "
            f"then rerun with --sif-cache {raw_path}"
        )
    with raw_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise RuntimeError(f"SIF cache must be a JSON object: {raw_path}")
    return payload


def parse_rank_change(change: Any) -> tuple[int | None, int | None]:
    if not change:
        return None, None
    parts = str(change).split("_")
    if len(parts) != 2:
        return None, None

    def one(text: str) -> int | None:
        stripped = text.strip()
        if not stripped or stripped.lower() == "null":
            return None
        try:
            return int(float(stripped))
        except ValueError:
            return None

    return one(parts[0]), one(parts[1])


def first_int(*values: Any) -> int | None:
    for value in values:
        parsed = as_int(value)
        if parsed is not None:
            return parsed
    return None


def extract_rank_info(detail: dict[str, Any], key: str) -> tuple[int | None, int | None]:
    pchange = detail.get("pchangeReason") or detail.get("changeReason") or {}
    info = pchange.get(key) or {}
    before, after = parse_rank_change(info.get("change"))
    before = before if before is not None else first_int(info.get("out"), info.get("before"), detail.get(f"{key}Before"))
    after = after if after is not None else first_int(info.get("in"), info.get("after"), detail.get(f"{key}After"))
    return before, after


def extract_sif_details(response: Any) -> list[dict[str, Any]]:
    if isinstance(response, list):
        return [item for item in response if isinstance(item, dict)]
    if not isinstance(response, dict):
        return []
    if isinstance(response.get("details"), list):
        return response["details"]
    data = response.get("data")
    if isinstance(data, dict):
        for key in ("details", "records", "list"):
            if isinstance(data.get(key), list):
                return data[key]
    for key in ("records", "list"):
        if isinstance(response.get(key), list):
            return response[key]
    return []


def normalize_sif(raw: dict[str, Any]) -> dict[date, dict[str, dict[str, Any]]]:
    daily = raw.get("daily") or raw.get("days") or raw.get("data") or {}
    if not isinstance(daily, dict):
        raise RuntimeError("SIF cache must contain a daily object keyed by date.")

    by_day: dict[date, dict[str, dict[str, Any]]] = {}
    for day_text, response in daily.items():
        day = parse_day(day_text)
        day_map: dict[str, dict[str, Any]] = {}
        for detail in extract_sif_details(response):
            keyword = str(detail.get("keyword") or detail.get("query") or "").strip().lower()
            if not keyword:
                continue
            nf_before, nf_after = extract_rank_info(detail, "nfInfo")
            sp_before, sp_after = extract_rank_info(detail, "spInfo")
            day_map[keyword] = {
                "keyword": keyword,
                "score": detail.get("score"),
                "score_ratio": detail.get("scoreRatio") or detail.get("score_ratio"),
                "diff_score": detail.get("diffScore"),
                "diff_score_ratio": detail.get("diffScoreRatio") or detail.get("diff_score_ratio"),
                "est_searches_num": detail.get("estSearchesNum") or detail.get("est_searches_num"),
                "searches_rank": detail.get("searchesRank") or detail.get("searches_rank"),
                "nf_rank_before": nf_before,
                "nf_rank_after": nf_after,
                "sp_rank_before": sp_before,
                "sp_rank_after": sp_after,
                "is_changed": detail.get("isChanged") or detail.get("is_changed"),
            }
        by_day[day] = day_map
    if not by_day:
        raise RuntimeError("SIF cache contains no usable daily keyword records.")
    return by_day


def median(values: list[float | int]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    return statistics.median(cleaned) if cleaned else None


def average(values: list[float | int]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    return statistics.mean(cleaned) if cleaned else None


def stable_baseline_days(days: list[date], launch_date: date | None) -> set[date]:
    ordered = sorted(days)
    if launch_date:
        selected = [day for day in ordered if day < launch_date]
    else:
        selected = ordered[: max(7, len(ordered) // 2)]
    return set(selected)


def select_core_keywords(sif_by_day: dict[date, dict[str, dict[str, Any]]], provided: list[str]) -> list[str]:
    if provided:
        return provided
    stats: dict[str, dict[str, list[float]]] = defaultdict(lambda: {"ratios": [], "scores": [], "ranks": []})
    for keyword_map in sif_by_day.values():
        for keyword, item in keyword_map.items():
            if item.get("nf_rank_after") is not None:
                stats[keyword]["ranks"].append(float(item["nf_rank_after"]))
            if isinstance(item.get("score_ratio"), (int, float)):
                stats[keyword]["ratios"].append(float(item["score_ratio"]))
            if isinstance(item.get("score"), (int, float)):
                stats[keyword]["scores"].append(float(item["score"]))
    ranked: list[tuple[str, float, float]] = []
    for keyword, values in stats.items():
        if not values["ranks"]:
            continue
        ranked.append((keyword, average(values["ratios"]) or 0, average(values["scores"]) or 0))
    ranked.sort(key=lambda item: (item[1], item[2]), reverse=True)
    return [keyword for keyword, _ratio, _score in ranked[:4]]


def select_stable_keywords(
    sif_by_day: dict[date, dict[str, dict[str, Any]]],
    launch_date: date | None,
    max_keywords: int = 30,
) -> list[str]:
    baseline = stable_baseline_days(list(sif_by_day.keys()), launch_date)
    min_days = min(7, max(3, len(baseline) // 2))
    stats: dict[str, dict[str, list[float]]] = defaultdict(lambda: {"ranks": [], "ratios": [], "scores": []})
    for day, keyword_map in sif_by_day.items():
        if day not in baseline:
            continue
        for keyword, item in keyword_map.items():
            rank = item.get("nf_rank_after")
            if rank is None:
                continue
            stats[keyword]["ranks"].append(float(rank))
            if isinstance(item.get("score_ratio"), (int, float)):
                stats[keyword]["ratios"].append(float(item["score_ratio"]))
            if isinstance(item.get("score"), (int, float)):
                stats[keyword]["scores"].append(float(item["score"]))
    selected: list[tuple[str, float, float]] = []
    for keyword, values in stats.items():
        med_rank = median(values["ranks"])
        avg_ratio = average(values["ratios"]) or 0
        if len(values["ranks"]) >= min_days and med_rank is not None and med_rank <= P1_MAX_RANK and avg_ratio >= 0.002:
            selected.append((keyword, med_rank, avg_ratio))
    selected.sort(key=lambda item: (item[1], -item[2], item[0]))
    return [keyword for keyword, _rank, _ratio in selected[:max_keywords]]


def keyword_prefix(keyword: str) -> str:
    safe = "".join(char if char.isalnum() else "_" for char in keyword.lower()).strip("_")
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe or "keyword"


def rank_down(start_rank: int | None, future_rank: int | None) -> bool:
    if start_rank is None:
        return False
    if start_rank <= P1_MAX_RANK and (future_rank is None or future_rank > P1_MAX_RANK):
        return True
    if future_rank is None:
        return False
    return future_rank - start_rank >= 5


def same_day_rank_down(item: dict[str, Any] | None) -> bool:
    if not item:
        return False
    return rank_down(item.get("nf_rank_before"), item.get("nf_rank_after"))


def build_panel(
    business_rows: list[BusinessRow],
    sif_by_day: dict[date, dict[str, dict[str, Any]]],
    core_keywords: list[str],
    stable_keywords: list[str],
    launch_date: date | None,
) -> list[dict[str, Any]]:
    panel: list[dict[str, Any]] = []
    for business in business_rows:
        keyword_map = sif_by_day.get(business.day, {})
        row: dict[str, Any] = {
            "date": business.day.isoformat(),
            "phase": "pre" if launch_date and business.day < launch_date else ("post" if launch_date else "all"),
            "sessions_total": business.sessions_total,
            "cvr": business.cvr,
            "ad_cvr": business.ad_cvr,
            "organic_cvr": business.organic_cvr,
            "clicks": business.clicks,
            "ad_orders": business.ad_orders,
            "organic_clicks": business.organic_clicks,
            "organic_orders": business.organic_orders,
            "units": business.units,
            "orders": business.orders,
            "organic_negative_adjustment": bool(
                (business.organic_orders is not None and business.organic_orders < 0)
                or (business.organic_cvr is not None and business.organic_cvr < 0)
            ),
            "sif_keyword_count": len(keyword_map),
        }
        for keyword in core_keywords:
            item = keyword_map.get(keyword)
            prefix = keyword_prefix(keyword)
            row[f"{prefix}_nf_rank"] = item.get("nf_rank_after") if item else None
            row[f"{prefix}_nf_rank_before"] = item.get("nf_rank_before") if item else None
            row[f"{prefix}_nf_same_day_down"] = same_day_rank_down(item)
            row[f"{prefix}_score_ratio"] = item.get("score_ratio") if item else None
        current_stable = [keyword for keyword in stable_keywords if keyword in keyword_map]
        stable_down_now = [keyword for keyword in current_stable if same_day_rank_down(keyword_map.get(keyword))]
        row["stable_keyword_present_count"] = len(current_stable)
        row["stable_keyword_same_day_down_count"] = len(stable_down_now)
        row["stable_keyword_same_day_down_ratio"] = len(stable_down_now) / len(current_stable) if current_stable else None
        panel.append(row)

    for lag in [0, 1, 2]:
        for index, row in enumerate(panel):
            future = panel[index + lag] if index + lag < len(panel) else None
            row[f"event_lag{lag}_evaluable"] = bool(future)
            core_down: list[str] = []
            stable_down_ratio = None
            if lag == 0:
                core_down = [keyword for keyword in core_keywords if row.get(f"{keyword_prefix(keyword)}_nf_same_day_down")]
                stable_down_ratio = row.get("stable_keyword_same_day_down_ratio")
            elif future:
                day = parse_day(row["date"])
                future_day = parse_day(future["date"])
                current_map = sif_by_day.get(day, {})
                future_map = sif_by_day.get(future_day, {})
                for keyword in core_keywords:
                    if rank_down(row.get(f"{keyword_prefix(keyword)}_nf_rank"), future.get(f"{keyword_prefix(keyword)}_nf_rank")):
                        core_down.append(keyword)
                stable_down = 0
                stable_compared = 0
                for keyword in stable_keywords:
                    current_item = current_map.get(keyword)
                    if not current_item:
                        continue
                    stable_compared += 1
                    future_item = future_map.get(keyword) or {}
                    if rank_down(current_item.get("nf_rank_after"), future_item.get("nf_rank_after")):
                        stable_down += 1
                stable_down_ratio = stable_down / stable_compared if stable_compared else None
            row[f"event_lag{lag}_core_down_count"] = len(core_down)
            row[f"event_lag{lag}_core_down_keywords"] = "; ".join(core_down)
            row[f"event_lag{lag}_stable_down_ratio"] = stable_down_ratio
            row[f"event_lag{lag}_rank_volatility"] = bool(
                row[f"event_lag{lag}_evaluable"]
                and (len(core_down) > 0 or (stable_down_ratio is not None and stable_down_ratio >= 0.30))
            )
    return panel


def percentile(values: list[float], p: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    pos = (len(ordered) - 1) * p
    lower = math.floor(pos)
    upper = math.ceil(pos)
    if lower == upper:
        return ordered[int(pos)]
    return ordered[lower] * (upper - pos) + ordered[upper] * (pos - lower)


def rolling_pre7_sessions(panel: list[dict[str, Any]], index: int) -> float | None:
    start = max(0, index - 7)
    values = [row.get("sessions_total") for row in panel[start:index] if isinstance(row.get("sessions_total"), (int, float))]
    return statistics.mean(values) if values else None


def threshold_candidates(panel: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for index, row in enumerate(panel):
        new_row = dict(row)
        prev_avg = rolling_pre7_sessions(panel, index)
        new_row["session_lift_vs_prev7"] = (
            (row["sessions_total"] / prev_avg - 1)
            if prev_avg and isinstance(row.get("sessions_total"), (int, float))
            else None
        )
        enriched.append(new_row)

    rows: list[dict[str, Any]] = []
    for lag in [0, 1, 2]:
        event_field = f"event_lag{lag}_rank_volatility"
        evaluable_field = f"event_lag{lag}_evaluable"
        evaluable = [row for row in enriched if row.get(evaluable_field)]
        event_count = sum(1 for row in evaluable if row.get(event_field))
        base_rate = event_count / len(evaluable) if evaluable else 0
        metric_specs = [
            ("cvr", "整体CVR", False),
            ("ad_cvr", "广告CVR", False),
            ("organic_cvr", "自然CVR", True),
        ]
        for metric, metric_name, exclude_negative_organic in metric_specs:
            metric_rows = [
                row
                for row in evaluable
                if isinstance(row.get(metric), (int, float))
                and not (exclude_negative_organic and row.get("organic_negative_adjustment"))
            ]
            values = sorted({float(row[metric]) for row in metric_rows})
            for threshold in values:
                triggered = [row for row in metric_rows if float(row[metric]) <= threshold]
                rows.append(build_candidate_row(triggered, event_count, event_field, base_rate, lag, metric, metric_name, threshold))

        cvr_rows = [row for row in evaluable if isinstance(row.get("cvr"), (int, float))]
        for threshold in sorted({float(row["cvr"]) for row in cvr_rows}):
            triggered = [
                row
                for row in cvr_rows
                if float(row["cvr"]) <= threshold
                and isinstance(row.get("session_lift_vs_prev7"), (int, float))
                and row["session_lift_vs_prev7"] >= 0.20
            ]
            if triggered:
                rows.append(
                    build_candidate_row(
                        triggered,
                        event_count,
                        event_field,
                        base_rate,
                        lag,
                        "traffic_expansion_low_cvr",
                        "Session放大+整体CVR",
                        threshold,
                    )
                )
    return rows


def build_candidate_row(
    triggered: list[dict[str, Any]],
    event_count: int,
    event_field: str,
    base_rate: float,
    lag: int,
    metric: str,
    metric_name: str,
    threshold: float,
) -> dict[str, Any]:
    true_positive = sum(1 for row in triggered if row.get(event_field))
    precision = true_positive / len(triggered) if triggered else 0
    recall = true_positive / event_count if event_count else 0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0
    return {
        "lag_days": lag,
        "metric": metric,
        "metric_name": metric_name,
        "threshold": threshold,
        "trigger_count": len(triggered),
        "event_count": event_count,
        "true_positive": true_positive,
        "false_positive": len(triggered) - true_positive,
        "false_negative": event_count - true_positive,
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "base_rate": base_rate,
        "lift": precision / base_rate if base_rate else None,
    }


def choose_thresholds(candidates: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, dict[str, Any] | None, dict[str, Any] | None]:
    viable = [row for row in candidates if row["trigger_count"] >= 4 and row["event_count"] >= 4]
    observation_pool = [
        row
        for row in viable
        if row["metric"] == "cvr"
        and row["recall"] >= 0.55
        and row["precision"] >= row["base_rate"]
        and (row["lift"] or 0) >= 1.15
    ]
    observation = max(
        observation_pool,
        key=lambda row: (row["f1"], row["recall"], row["precision"], -row["threshold"]),
        default=None,
    )
    danger_pool = [
        row
        for row in viable
        if row["metric"] == "cvr"
        and row["trigger_count"] >= 4
        and row["precision"] >= max(0.50, row["base_rate"] * 1.25)
        and (row["lift"] or 0) >= 1.20
    ]
    danger = max(danger_pool, key=lambda row: (row["f1"], row["precision"], row["lift"] or 0, row["trigger_count"]), default=None)
    ad_pool = [
        row
        for row in viable
        if row["metric"] == "ad_cvr"
        and row["trigger_count"] >= 4
        and row["precision"] >= row["base_rate"] * 1.20
        and (row["lift"] or 0) >= 1.20
    ]
    ad_confirmation = max(ad_pool, key=lambda row: (row["f1"], row["precision"], row["lift"] or 0), default=None)
    return observation, danger, ad_confirmation


def apply_threshold_flags(
    panel: list[dict[str, Any]],
    observation: dict[str, Any] | None,
    danger: dict[str, Any] | None,
    ad_confirmation: dict[str, Any] | None,
) -> None:
    for row in panel:
        row["trigger_observation_line"] = is_triggered(row, observation)
        row["trigger_danger_line"] = is_triggered(row, danger)
        row["trigger_ad_cvr_confirmation_line"] = is_triggered(row, ad_confirmation)


def is_triggered(row: dict[str, Any], threshold: dict[str, Any] | None) -> bool:
    if not threshold:
        return False
    metric = threshold["metric"]
    if metric == "traffic_expansion_low_cvr":
        return (
            isinstance(row.get("cvr"), (int, float))
            and row["cvr"] <= threshold["threshold"]
            and isinstance(row.get("session_lift_vs_prev7"), (int, float))
            and row["session_lift_vs_prev7"] >= 0.20
        )
    return isinstance(row.get(metric), (int, float)) and row[metric] <= threshold["threshold"]


def summarize_metric(panel: list[dict[str, Any]], field: str, exclude_negative_organic: bool = False) -> dict[str, float | None]:
    values = [
        float(row[field])
        for row in panel
        if isinstance(row.get(field), (int, float)) and not (exclude_negative_organic and row.get("organic_negative_adjustment"))
    ]
    return {
        "min": min(values) if values else None,
        "p25": percentile(values, 0.25),
        "median": percentile(values, 0.50),
        "p75": percentile(values, 0.75),
        "max": max(values) if values else None,
    }


def event_summary(panel: list[dict[str, Any]], lag: int) -> dict[str, Any]:
    field = f"event_lag{lag}_rank_volatility"
    evaluable = [row for row in panel if row.get(f"event_lag{lag}_evaluable")]
    events = [row for row in evaluable if row.get(field)]
    return {
        "lag": lag,
        "evaluable_days": len(evaluable),
        "days": len(events),
        "rate": len(events) / len(evaluable) if evaluable else 0,
        "dates": [row["date"] for row in events],
    }


def top_rows(rows: list[dict[str, Any]], metric: str, limit: int = 6) -> list[dict[str, Any]]:
    subset = [row for row in rows if row["metric"] == metric and row["trigger_count"] >= 4]
    subset.sort(key=lambda row: (row["f1"], row["precision"], row["recall"]), reverse=True)
    return subset[:limit]


def threshold_line(row: dict[str, Any] | None) -> str:
    if not row:
        return "未找到足够稳健的阈值"
    return (
        f"{row['metric_name']} <= {fmt_pct(row['threshold'])} "
        f"(滞后 {row['lag_days']} 天，触发 {row['trigger_count']} 天，命中率 {fmt_pct(row['precision'])}，"
        f"召回率 {fmt_pct(row['recall'])}，Lift {fmt_num(row.get('lift'), 2)})"
    )


def candidate_table(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "无可用候选。"
    lines = ["| 指标 | 滞后 | 阈值 | 触发天数 | 命中天数 | 命中率 | 召回率 | Lift | F1 |", "|---|---:|---:|---:|---:|---:|---:|---:|---:|"]
    for row in rows:
        lines.append(
            "| {metric_name} | {lag_days} | {threshold} | {trigger_count} | {tp} | {precision} | {recall} | {lift} | {f1} |".format(
                metric_name=row["metric_name"],
                lag_days=row["lag_days"],
                threshold=fmt_pct(row["threshold"]),
                trigger_count=row["trigger_count"],
                tp=row["true_positive"],
                precision=fmt_pct(row["precision"]),
                recall=fmt_pct(row["recall"]),
                lift=fmt_num(row.get("lift"), 2),
                f1=fmt_num(row["f1"], 2),
            )
        )
    return "\n".join(lines)


def metric_summary_table(panel: list[dict[str, Any]]) -> str:
    specs = [
        ("整体CVR", "cvr", False),
        ("广告CVR", "ad_cvr", False),
        ("自然CVR", "organic_cvr", True),
    ]
    lines = ["| 指标 | Min | P25 | Median | P75 | Max |", "|---|---:|---:|---:|---:|---:|"]
    for name, field, exclude_negative in specs:
        stats = summarize_metric(panel, field, exclude_negative)
        lines.append(
            f"| {name} | {fmt_pct(stats['min'])} | {fmt_pct(stats['p25'])} | {fmt_pct(stats['median'])} | {fmt_pct(stats['p75'])} | {fmt_pct(stats['max'])} |"
        )
    return "\n".join(lines)


def event_summary_table(panel: list[dict[str, Any]]) -> str:
    lines = ["| 滞后窗口 | 可评估天数 | 波动天数 | 波动率 | 日期 |", "|---|---:|---:|---:|---|"]
    for item in [event_summary(panel, lag) for lag in [0, 1, 2]]:
        dates = ", ".join(item["dates"][:12])
        if len(item["dates"]) > 12:
            dates += "..."
        lines.append(f"| {item['lag']} 天 | {item['evaluable_days']} | {item['days']} | {fmt_pct(item['rate'])} | {dates} |")
    return "\n".join(lines)


def make_report(
    config: AnalysisConfig,
    panel: list[dict[str, Any]],
    core_keywords: list[str],
    stable_keywords: list[str],
    candidates: list[dict[str, Any]],
    observation: dict[str, Any] | None,
    danger: dict[str, Any] | None,
    ad_confirmation: dict[str, Any] | None,
    output_paths: dict[str, Path],
    raw_sif: dict[str, Any],
) -> str:
    negative_days = [row["date"] for row in panel if row.get("organic_negative_adjustment")]
    main_metric = observation["metric_name"] if observation else (danger["metric_name"] if danger else "未形成单一主指标")
    missing_sif_days = [row["date"] for row in panel if not row.get("sif_keyword_count")]
    status = "DONE_WITH_CONCERNS" if missing_sif_days or not observation or not danger else "DONE"
    launch_text = config.launch_date.isoformat() if config.launch_date else "未提供"
    core_rank_headers = ["日期", "Session", "CVR", "广告CVR", "自然CVR"] + core_keywords[:4] + ["观察线", "危险线", "广告确认线"]
    core_lines = ["| " + " | ".join(core_rank_headers) + " |", "|" + "|".join(["---"] + ["---:"] * (len(core_rank_headers) - 1)) + "|"]
    for row in panel:
        ranks = [str(row.get(f"{keyword_prefix(keyword)}_nf_rank") or "") for keyword in core_keywords[:4]]
        cells = [
            row["date"],
            str(row.get("sessions_total") or ""),
            fmt_pct(row.get("cvr")),
            fmt_pct(row.get("ad_cvr")),
            fmt_pct(row.get("organic_cvr")),
            *ranks,
            "Y" if row.get("trigger_observation_line") else "",
            "Y" if row.get("trigger_danger_line") else "",
            "Y" if row.get("trigger_ad_cvr_confirmation_line") else "",
        ]
        core_lines.append("| " + " | ".join(cells) + " |")

    concerns = []
    if missing_sif_days:
        concerns.append(f"SIF 关键词数据缺失 {len(missing_sif_days)} 天")
    if len(panel) < 21:
        concerns.append("样本少于 21 天，只适合验证流程或做轻量参考")
    if not observation:
        concerns.append("观察线未达到最小稳健条件")
    if not danger:
        concerns.append("危险线未达到最小稳健条件")
    if not ad_confirmation:
        concerns.append("广告 CVR 确认线本次未达到强确认条件")

    raw_source = raw_sif.get("source") or "user-provided SIF cache"
    return f"""---
created: {datetime.now().strftime('%Y-%m-%d %H:%M')}
topic: Amazon {config.site} ASIN {config.asin} CVR 自然排名阈值分析
type: CVR自然排名阈值分析报告
brand: {config.brand}
site: {config.site}
asin: {config.asin}
data_sources:
  - business_report: {config.business_report.name}
  - sif_cache: {raw_source}
window: {panel[0]['date']} to {panel[-1]['date']}
launch_start: {launch_text}
status: {status}
---

# {config.asin} CVR 自然排名阈值分析

## 结论

本窗口内最适合作为主监控的指标是 **{main_metric}**。

- 观察线：**{threshold_line(observation)}**
- 危险线：**{threshold_line(danger)}**
- 广告 CVR 确认线：**{threshold_line(ad_confirmation)}**

这些线是运营预警线，不是 Amazon 官方算法阈值。它们回答的是：当业务报告里的 CVR 低到某个区间时，SIF 监控到的核心词 / 稳定词自然排名波动概率是否明显变高。

## 阈值评估指标怎么理解

### 命中率

运营含义：这个阈值报警的时候，有多大概率真的会出现排名波动。

例子：如果危险线触发了 17 天，其中 9 天之后核心词真的掉排名，那命中率就是 `9 / 17 = 52.94%`。

怎么用：命中率越高，越适合做“暂停/收缩放量”的判断，因为它代表误报更少。

### 召回率

运营含义：所有真实发生排名波动的日期里，这条阈值提前抓到了多少。

例子：如果一共发生 15 次核心词波动，其中 9 次在波动前已经触发 CVR 危险线，那召回率就是 `9 / 15 = 60%`。

怎么用：召回率越高，越适合做“观察线”，因为它代表漏掉的风险更少。

### Lift

运营含义：触发阈值以后，排名波动概率比平时高了多少倍。

例子：平时核心词波动概率是 32%，低于某条 CVR 阈值后波动概率变成 53%，那 Lift 约为 `53% / 32% = 1.66`。

怎么用：Lift 用来判断这条阈值是不是“真的有区分度”。如果 Lift 接近 1，说明低于这个 CVR 和排名波动没什么明显关系。

这三个指标不是 SIF 原生字段。SIF 提供的是关键词、日期、自然排名、广告排名和流量信号；本 Skill 会自己定义“排名波动事件”，再用业务报告里的 CVR 回测这些事件。所以最终阈值只作为运营预警线，不单独决定预算动作。

## 数据事实

- 业务报告：{len(panel)} 天，{panel[0]['date']} 至 {panel[-1]['date']}，日期连续。
- SIF 有效日：{sum(1 for row in panel if row.get('sif_keyword_count'))} 天。
- 核心词：{', '.join(core_keywords) if core_keywords else '未能自动筛出核心词'}。
- 稳定词篮子：{len(stable_keywords)} 个，示例：{', '.join(stable_keywords[:12]) if stable_keywords else '无'}。
- 自然订单 / 自然 CVR 异常调整日：{', '.join(negative_days) if negative_days else '无'}。
- 运行关注事项：{'; '.join(concerns) if concerns else '无'}。

### CVR 分布

{metric_summary_table(panel)}

### 排名波动事件

{event_summary_table(panel)}

## 阈值候选

### 整体 CVR

{candidate_table(top_rows(candidates, 'cvr'))}

### 广告 CVR

{candidate_table(top_rows(candidates, 'ad_cvr'))}

### 自然 CVR

{candidate_table(top_rows(candidates, 'organic_cvr'))}

### Session 放大 + 低 CVR

{candidate_table(top_rows(candidates, 'traffic_expansion_low_cvr'))}

## 日级观察面板

{chr(10).join(core_lines)}

## 运营建议

1. 站外放量期间，每天先看观察线；连续 2 天触发时，不建议继续放大冷流量。
2. 触发危险线时，优先暂停新增站外预算，检查素材承诺、价格、Coupon、配送和 Listing 首屏承接。
3. 如果 Session 明显放大但整体 CVR 同时跌破观察线，应按“低意图流量稀释”处理，而不是只看广告订单有没有增长。
4. 自然 CVR 若存在负订单调整日，不作为主监控指标；它只能辅助解释。

## 附件

- 日级面板 CSV：`{output_paths['panel'].name}`
- 阈值候选 CSV：`{output_paths['candidates'].name}`
- SIF 原始缓存：`{output_paths['raw_sif'].name}`
- 稳定词篮子 JSON：`{output_paths['stable_keywords'].name}`
"""


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_outputs(
    config: AnalysisConfig,
    panel: list[dict[str, Any]],
    stable_keywords: list[str],
    candidates: list[dict[str, Any]],
    report: str,
    raw_sif: dict[str, Any],
) -> dict[str, Path]:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    base = f"{config.analysis_date}_{config.site}_{config.asin}"
    paths = {
        "report": config.output_dir / f"{base}_CVR自然排名阈值分析.md",
        "panel": config.output_dir / f"{base}_CVR排名阈值_日级面板.csv",
        "candidates": config.output_dir / f"{base}_CVR排名阈值_候选阈值.csv",
        "stable_keywords": config.output_dir / f"{base}_CVR排名阈值_稳定词篮子.json",
        "raw_sif": config.sif_cache or default_sif_cache_path(config),
    }
    write_csv(paths["panel"], panel)
    write_csv(paths["candidates"], candidates)
    paths["stable_keywords"].write_text(json.dumps(stable_keywords, ensure_ascii=False, indent=2), encoding="utf-8")
    if config.sif_cache is None:
        paths["raw_sif"].write_text(json.dumps(raw_sif, ensure_ascii=False, indent=2), encoding="utf-8")
    paths["report"].write_text(report, encoding="utf-8")
    return paths


def run_analysis(config: AnalysisConfig) -> dict[str, Any]:
    business_rows = load_business_rows(config)
    raw_sif = load_sif_cache(config)
    sif_by_day = normalize_sif(raw_sif)
    core_keywords = select_core_keywords(sif_by_day, config.core_keywords)
    stable_keywords = select_stable_keywords(sif_by_day, config.launch_date)
    panel = build_panel(business_rows, sif_by_day, core_keywords, stable_keywords, config.launch_date)
    candidates = threshold_candidates(panel)
    observation, danger, ad_confirmation = choose_thresholds(candidates)
    apply_threshold_flags(panel, observation, danger, ad_confirmation)

    placeholder_paths = {
        "panel": config.output_dir / f"{config.analysis_date}_{config.site}_{config.asin}_CVR排名阈值_日级面板.csv",
        "candidates": config.output_dir / f"{config.analysis_date}_{config.site}_{config.asin}_CVR排名阈值_候选阈值.csv",
        "raw_sif": config.sif_cache or default_sif_cache_path(config),
        "stable_keywords": config.output_dir / f"{config.analysis_date}_{config.site}_{config.asin}_CVR排名阈值_稳定词篮子.json",
    }
    report = make_report(
        config,
        panel,
        core_keywords,
        stable_keywords,
        candidates,
        observation,
        danger,
        ad_confirmation,
        placeholder_paths,
        raw_sif,
    )
    paths = write_outputs(config, panel, stable_keywords, candidates, report, raw_sif)
    return {
        "business_rows": len(business_rows),
        "sif_days": len(sif_by_day),
        "core_keywords": core_keywords,
        "stable_keywords": stable_keywords,
        "observation": observation,
        "danger": danger,
        "ad_confirmation": ad_confirmation,
        "paths": {key: str(value) for key, value in paths.items()},
    }


def main() -> None:
    config = build_config(parse_args())
    result = run_analysis(config)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
